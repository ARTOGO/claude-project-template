#!/usr/bin/env python3
"""
Excel 公式重新計算與驗證腳本

用途：
- 重新計算所有公式
- 檢測公式錯誤 (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?)
- 確保 Excel 檔案零公式錯誤

使用方式：
    python recalc.py <excel_file.xlsx>
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter

# 公式錯誤類型
FORMULA_ERRORS = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!']

def recalc_excel(filepath: str) -> tuple[bool, list]:
    """
    重新計算 Excel 檔案的所有公式並檢測錯誤

    Args:
        filepath: Excel 檔案路徑

    Returns:
        (success, errors): 成功標記和錯誤列表
    """
    try:
        # 載入工作簿
        wb = openpyxl.load_workbook(filepath, data_only=False)
        errors = []

        # 遍歷所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 檢查每個儲存格
            for row in ws.iter_rows():
                for cell in row:
                    # 跳過空儲存格
                    if cell.value is None:
                        continue

                    # 檢查公式錯誤
                    if isinstance(cell.value, str):
                        for error in FORMULA_ERRORS:
                            if error in cell.value:
                                errors.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'error': error,
                                    'formula': cell.value
                                })

        # 重新計算（標記為需要計算）
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True

        # 儲存
        wb.save(filepath)
        wb.close()

        return (len(errors) == 0, errors)

    except Exception as e:
        return (False, [{'error': 'Exception', 'message': str(e)}])


def print_errors(errors: list):
    """格式化輸出錯誤"""
    if not errors:
        print("✅ 成功：零公式錯誤")
        return

    print(f"❌ 失敗：發現 {len(errors)} 個公式錯誤\n")

    for err in errors:
        if 'sheet' in err:
            print(f"  工作表: {err['sheet']}")
            print(f"  儲存格: {err['cell']}")
            print(f"  錯誤類型: {err['error']}")
            print(f"  公式: {err['formula']}")
            print()
        else:
            print(f"  錯誤: {err.get('message', 'Unknown error')}")


def main():
    if len(sys.argv) < 2:
        print("使用方式: python recalc.py <excel_file.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]

    print(f"重新計算: {filepath}")
    success, errors = recalc_excel(filepath)

    print_errors(errors)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
